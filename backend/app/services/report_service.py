from sqlalchemy.orm import Session
from app.models.product import Product, Offer, PriceHistory, Seller
from app.services.product_service import ProductService
from app.services.analytics import AnalyticsService
from app.services.ai_service import AIService
from app.core.minio_client import minio_client
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import io
import logging

logger = logging.getLogger(__name__)


class ReportService:
    @staticmethod
    def generate_product_excel(db: Session, product_id: int) -> str:
        product = ProductService.get_product(db, product_id)
        if not product:
            raise ValueError("Product not found")
        
        wb = Workbook()
        
        ws1 = wb.active
        ws1.title = "Текущие предложения"
        
        headers = ["Позиция", "Продавец", "Цена", "Рейтинг", "Отзывы", "В наличии"]
        ws1.append(headers)
        
        for offer in sorted(product.offers, key=lambda x: x.price):
            ws1.append([
                offer.position or "-",
                offer.seller.name,
                offer.price,
                offer.seller.rating or "-",
                offer.seller.reviews_count,
                "Да" if offer.in_stock else "Нет"
            ])
        
        ws2 = wb.create_sheet("История цен")
        ws2.append(["Дата", "Продавец", "Цена", "Позиция"])
        
        cutoff_date = datetime.utcnow() - timedelta(days=30)
        history = db.query(PriceHistory).filter(
            PriceHistory.product_id == product_id,
            PriceHistory.recorded_at >= cutoff_date
        ).order_by(PriceHistory.recorded_at).all()
        
        for record in history:
            ws2.append([
                record.recorded_at.strftime("%Y-%m-%d"),
                record.seller.name if record.seller else "-",
                record.price,
                record.position or "-"
            ])
        
        ws3 = wb.create_sheet("Статистика")
        prices = [o.price for o in product.offers]
        
        if prices:
            ws3.append(["Метрика", "Значение"])
            ws3.append(["Минимальная цена", min(prices)])
            ws3.append(["Максимальная цена", max(prices)])
            ws3.append(["Средняя цена", sum(prices) / len(prices)])
            ws3.append(["Количество предложений", len(prices)])
        
        try:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            object_name = f"reports/product_{product_id}_{timestamp}.xlsx"
            
            file_data = buffer.read()
            if not file_data:
                raise ValueError("Generated Excel file is empty")
            
            minio_client.upload_bytes(
                file_data,
                object_name,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            return object_name
        except Exception as e:
            logger.error(f"Error saving report to MinIO: {e}", exc_info=True)
            raise
    
    @staticmethod
    def generate_comparison_excel(db: Session, product_id_1: int, product_id_2: int) -> str:
        product1 = ProductService.get_product(db, product_id_1)
        product2 = ProductService.get_product(db, product_id_2)
        
        if not product1 or not product2:
            raise ValueError("One or both products not found")
        
        wb = Workbook()
        
        ws1 = wb.active
        ws1.title = "Сравнение"
        
        ws1.append(["Метрика", product1.name[:30] if product1.name else "Товар 1", product2.name[:30] if product2.name else "Товар 2"])
        
        prices1 = [o.price for o in product1.offers]
        prices2 = [o.price for o in product2.offers]
        
        if prices1 and prices2:
            ws1.append(["Минимальная цена", min(prices1), min(prices2)])
            ws1.append(["Максимальная цена", max(prices1), max(prices2)])
            ws1.append(["Средняя цена", sum(prices1) / len(prices1), sum(prices2) / len(prices2)])
            ws1.append(["Количество предложений", len(prices1), len(prices2)])
        
        ws2 = wb.create_sheet(f"{product1.name[:30] if product1.name else 'Товар 1'}")
        ws2.append(["Позиция", "Продавец", "Цена", "Рейтинг"])
        for offer in sorted(product1.offers, key=lambda x: x.price):
            ws2.append([
                offer.position or "-",
                offer.seller.name,
                offer.price,
                offer.seller.rating or "-"
            ])
        
        ws3 = wb.create_sheet(f"{product2.name[:30] if product2.name else 'Товар 2'}")
        ws3.append(["Позиция", "Продавец", "Цена", "Рейтинг"])
        for offer in sorted(product2.offers, key=lambda x: x.price):
            ws3.append([
                offer.position or "-",
                offer.seller.name,
                offer.price,
                offer.seller.rating or "-"
            ])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        object_name = f"reports/comparison_{product_id_1}_vs_{product_id_2}_{timestamp}.xlsx"
        
        minio_client.upload_bytes(
            buffer.read(),
            object_name,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        return object_name

    @staticmethod
    def generate_advanced_analytics_report(db: Session, product_id: int, user_price: float = None) -> str:
        product = ProductService.get_product(db, product_id)
        if not product:
            raise ValueError("Product not found")
        
        offers_data = [
            {
                "price": o.price,
                "seller_name": o.seller.name,
                "seller_rating": o.seller.rating,
                "seller_reviews_count": o.seller.reviews_count,
                "position": o.position
            }
            for o in product.offers
        ]
        
        price_history_data = []
        cutoff_date = datetime.utcnow() - timedelta(days=90)
        history = db.query(PriceHistory).filter(
            PriceHistory.product_id == product_id,
            PriceHistory.recorded_at >= cutoff_date
        ).order_by(PriceHistory.recorded_at).all()
        
        for record in history:
            price_history_data.append({
                "date": record.recorded_at.date().isoformat(),
                "price": record.price,
                "position": record.position,
                "seller_name": record.seller.name if record.seller else "Unknown"
            })
        
        price_dist = AnalyticsService.calculate_price_distribution(offers_data)
        volatility = AnalyticsService.calculate_volatility(price_history_data)
        trend = AnalyticsService.detect_trend(price_history_data)
        demand_proxy = AnalyticsService.calculate_demand_proxy(offers_data, price_history_data)
        entry_barrier = AnalyticsService.calculate_entry_barrier(offers_data)
        optimal_price = AnalyticsService.calculate_optimal_price(offers_data)
        anomalies = AnalyticsService.detect_anomalies(price_history_data, offers_data)
        weighted_rank = AnalyticsService.calculate_weighted_rank(offers_data, user_price)
        dominant_sellers = AnalyticsService.detect_dominant_sellers(offers_data, price_history_data)
        
        ai_service = AIService()
        ai_insights = ai_service.generate_advanced_insights(
            product.name or f"Товар {product.kaspi_id}",
            price_dist,
            volatility,
            trend,
            demand_proxy,
            entry_barrier,
            optimal_price,
            anomalies,
            weighted_rank,
            dominant_sellers,
            user_price
        )
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Расширенная аналитика"
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = f"РАСШИРЕННАЯ АНАЛИТИКА: {product.name or f'Товар {product.kaspi_id}'}"
        cell.font = Font(bold=True, size=14)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 2
        
        ws[f'A{row}'] = "РАСПРЕДЕЛЕНИЕ ЦЕН"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        row += 1
        
        metrics = [
            ("Минимальная цена", price_dist.get("min")),
            ("Максимальная цена", price_dist.get("max")),
            ("Медианная цена", price_dist.get("median")),
            ("P25 (25-й перцентиль)", price_dist.get("p25")),
            ("P75 (75-й перцентиль)", price_dist.get("p75")),
            ("IQR (межквартильный размах)", price_dist.get("iqr")),
            ("Средняя цена", price_dist.get("mean")),
            ("Стандартное отклонение", price_dist.get("std"))
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = f"{value:.2f}" if value is not None else "N/A"
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        row += 1
        ws[f'A{row}'] = "ВОЛАТИЛЬНОСТЬ И ТРЕНДЫ"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        row += 1
        
        volatility_metrics = [
            ("Волатильность (σ)", volatility.get("volatility")),
            ("Коэффициент вариации", volatility.get("coefficient_of_variation")),
            ("Диапазон цен", volatility.get("price_range")),
            ("Направление тренда", trend.get("direction", "N/A")),
            ("Изменение за период", f"{trend.get('change_percent', 0):.2f}%"),
            ("SMA (простая скользящая)", trend.get("sma")),
            ("EMA (экспоненциальная)", trend.get("ema"))
        ]
        
        for label, value in volatility_metrics:
            ws[f'A{row}'] = label
            if isinstance(value, (int, float)):
                ws[f'B{row}'] = f"{value:.2f}" if value is not None else "N/A"
            else:
                ws[f'B{row}'] = value
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        row += 1
        ws[f'A{row}'] = "АНАЛИЗ СПРОСА И КОНКУРЕНЦИИ"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        row += 1
        
        demand_metrics = [
            ("Оценка спроса", demand_proxy.get("demand_score")),
            ("Количество продавцов", demand_proxy.get("sellers_count")),
            ("Уровень конкуренции", demand_proxy.get("competition_level")),
            ("Средний рейтинг", demand_proxy.get("avg_rating")),
            ("Барьер входа", entry_barrier.get("level")),
            ("Оценка барьера", entry_barrier.get("barrier_score"))
        ]
        
        for label, value in demand_metrics:
            ws[f'A{row}'] = label
            if isinstance(value, (int, float)):
                ws[f'B{row}'] = f"{value:.2f}" if value is not None else "N/A"
            else:
                ws[f'B{row}'] = value
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        if entry_barrier.get("factors"):
            row += 1
            ws[f'A{row}'] = "Факторы барьера входа:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            for factor in entry_barrier.get("factors", []):
                ws[f'A{row}'] = f"  • {factor}"
                row += 1
        
        row += 1
        ws[f'A{row}'] = "ОПТИМАЛЬНАЯ ЦЕНА"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        row += 1
        
        optimal_metrics = [
            ("Рекомендуемая цена", optimal_price.get("optimal_price")),
            ("Ожидаемая позиция", optimal_price.get("estimated_position")),
            ("Маржа (%)", optimal_price.get("margin_percent")),
            ("Маржа (сумма)", optimal_price.get("margin_amount")),
            ("Себестоимость", optimal_price.get("cost_price"))
        ]
        
        for label, value in optimal_metrics:
            ws[f'A{row}'] = label
            if isinstance(value, (int, float)):
                ws[f'B{row}'] = f"{value:.2f}" if value is not None else "N/A"
            else:
                ws[f'B{row}'] = value
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        if anomalies:
            row += 1
            ws[f'A{row}'] = "ОБНАРУЖЕННЫЕ АНОМАЛИИ"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            row += 1
            
            for anomaly in anomalies[:5]:
                ws[f'A{row}'] = anomaly.get("message", "")
                ws[f'A{row}'].font = Font(color="FF0000")
                row += 1
        
        if dominant_sellers:
            row += 1
            ws[f'A{row}'] = "ДОМИНИРУЮЩИЕ ПРОДАВЦЫ"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            row += 1
            
            headers = ["Продавец", "Частота в TOP-3", "Средняя позиция", "Оценка доминирования"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = section_fill
                cell.border = border
            row += 1
            
            for seller in dominant_sellers[:10]:
                ws[f'A{row}'] = seller.get("seller_name", "")
                ws[f'B{row}'] = seller.get("top3_frequency", 0)
                ws[f'C{row}'] = f"{seller.get('avg_position', 0):.1f}"
                ws[f'D{row}'] = f"{seller.get('dominance_score', 0):.2f}"
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = border
                row += 1
        
        row += 2
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "AI-ГЕНЕРИРОВАННЫЕ ИНСАЙТЫ И РЕКОМЕНДАЦИИ"
        cell.font = Font(bold=True, size=12)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = ai_insights
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        object_name = f"reports/advanced_analytics_{product_id}_{timestamp}.xlsx"
        
        minio_client.upload_bytes(
            buffer.read(),
            object_name,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        return object_name
